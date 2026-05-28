# -*- coding: utf-8 -*-
"""
Bilibili 评论爬取工具（视频 / 动态，主评论 + 楼中楼，游标翻页无上限）

功能：
  - 支持视频评论区、动态评论区（转发类 type=17 / 图文类 type=11）
  - 抓取全量主评论（游标接口，突破普通分页 5000 条上限）
  - 可选抓取楼中楼（二级评论），按"回复数阈值"过滤，只抓引起讨论的评论
  - 表格标注每条是「主评论」还是「楼中楼」
  - 边抓边写入 SQLite，支持断点续传、Ctrl+C 主动停止
  - 写入 Excel 前过滤非法字符，避免 openpyxl 报错

依赖：
    pip install bilibili-api-python openpyxl httpx

打包成 exe（在 Windows 上执行）：
    pip install pyinstaller
    pyinstaller -F -n bilibili_comment_tool ^
      --collect-all bilibili_api ^
      --collect-all httpx ^
      --collect-all httpcore ^
      --collect-all anyio ^
      find_user_comments.py
"""

import asyncio
import httpx
import hashlib
import time
import re
import sqlite3
import os
from collections import defaultdict

from bilibili_api import video, Credential

# ============ 安全下限 ============
最小请求间隔秒 = 0.5
默认请求间隔秒 = 1.0
默认最大条数   = 0        # 0 = 不限制，抓到底
默认回复阈值   = 5        # 回复数 >= 此值的主评论才抓楼中楼
# =================================

MIXIN_KEY_ENC_TAB = [
    46,47,18,2,53,8,23,32,15,50,10,31,58,3,45,35,
    27,43,5,49,33,9,42,19,29,28,14,39,12,38,41,13,
    37,48,7,16,24,55,40,61,26,17,0,1,60,51,30,4,
    22,25,54,21,56,59,6,63,57,62,11,36,20,34,44,52,
]

_非法字符正则 = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f"
    r"\ud800-\udfff"
    r"\ufffe\uffff]"
)

def 清理非法字符(text: str) -> str:
    if not isinstance(text, str):
        return str(text)
    return _非法字符正则.sub("", text)


# ── 交互输入 ────────────────────────────────────────────────────────

def 询问(提示文字, 默认值=None):
    后缀 = f"（默认 {默认值}）" if 默认值 is not None else ""
    答 = input(f"{提示文字}{后缀}：").strip()
    if not 答 and 默认值 is not None:
        return str(默认值)
    return 答


def 收集配置():
    print("=" * 50)
    print("   Bilibili 评论统计工具（主评论 + 楼中楼）")
    print("=" * 50)
    print("提示：cookie 从浏览器获取 —— 登录B站 → F12 →")
    print("      Application → Cookies → https://www.bilibili.com")
    print("      建议使用小号，cookie 等同登录凭证，请勿外传。")
    print("-" * 50)

    sessdata = 询问("请输入 SESSDATA")
    while not sessdata:
        print("  SESSDATA 不能为空，请重新输入。")
        sessdata = 询问("请输入 SESSDATA")

    bili_jct = 询问("请输入 bili_jct")
    buvid3   = 询问("请输入 buvid3")

    # 评论区类型
    print("-" * 50)
    print("请选择要抓取的评论区类型：")
    print("  1 = 视频评论区（输入 BV 号）")
    print("  2 = 动态评论区 - 转发/纯文字类（输入动态 ID）")
    print("  3 = 动态评论区 - 图文相册类（输入动态 ID）")
    print("  提示：动态 ID 是动态链接 t.bilibili.com/ 后面那串数字。")
    print("        若类型 2 抓不到评论，请改用类型 3，反之亦然。")
    while True:
        类型选择 = 询问("请选择（1 / 2 / 3）", "1")
        if 类型选择 in ("1", "2", "3"):
            break
        print("  请输入 1、2 或 3。")

    bvid = ""
    oid_直填 = None
    if 类型选择 == "1":
        评论类型 = 1
        bvid = 询问("请输入目标视频 BV 号（如 BV1fSGJ69EbE）")
        while not bvid:
            print("  BV 号不能为空，请重新输入。")
            bvid = 询问("请输入目标视频 BV 号")
        目标标识 = bvid
    else:
        评论类型 = 17 if 类型选择 == "2" else 11
        while True:
            动态id = 询问("请输入动态 ID（链接 t.bilibili.com/ 后的数字串）").strip()
            if 动态id.isdigit():
                oid_直填 = int(动态id)
                break
            print("  动态 ID 应该是一串纯数字，请重新输入。")
        目标标识 = f"dyn{oid_直填}"

    # 是否抓楼中楼
    print("-" * 50)
    print("是否抓取楼中楼（二级评论）？")
    print("  楼中楼能让统计更完整（刷屏常藏在回复里），但会增加抓取时间。")
    while True:
        楼中楼选择 = 询问("抓取楼中楼？（y / n）", "y").lower()
        if 楼中楼选择 in ("y", "n"):
            抓楼中楼 = (楼中楼选择 == "y")
            break
        print("  请输入 y 或 n。")

    # 阈值在抓完主评论后再问（那时能给出分布参考），此处先存默认
    回复阈值 = 默认回复阈值

    while True:
        原始 = 询问("最多抓取多少条主评论（0 表示不限制，抓到底）", 默认最大条数)
        try:
            最大条数 = int(原始)
            break
        except ValueError:
            print("  请输入一个整数。")

    while True:
        原始 = 询问(f"请输入请求延迟秒数（不低于 {最小请求间隔秒}）", 默认请求间隔秒)
        try:
            延迟 = float(原始)
        except ValueError:
            print("  请输入一个数字。")
            continue
        if 延迟 < 最小请求间隔秒:
            print(f"  延迟过低，已自动提升到安全下限 {最小请求间隔秒} 秒。")
            延迟 = 最小请求间隔秒
        break

    print("-" * 50)
    print("输出模式：")
    print("  1 = 完整保留每条评论内容（默认）")
    print("  2 = 仅统计评论数量（不保留内容，表格更精简）")
    while True:
        模式 = 询问("请选择输出模式（1 或 2）", "1")
        if 模式 in ("1", "2"):
            保留内容 = (模式 == "1")
            break
        print("  请输入 1 或 2。")

    输出文件名 = 询问("请输入导出文件名（不含扩展名）", "B站评论统计结果")

    print("-" * 50)
    return {
        "凭证": Credential(sessdata=sessdata, bili_jct=bili_jct, buvid3=buvid3),
        "评论类型": 评论类型,
        "bvid": bvid,
        "oid_直填": oid_直填,
        "目标标识": 目标标识,
        "抓楼中楼": 抓楼中楼,
        "回复阈值": 回复阈值,
        "最大条数": 最大条数,
        "请求间隔秒": 延迟,
        "保留内容": 保留内容,
        "输出文件名": 输出文件名,
    }


# ── SQLite 落盘 ─────────────────────────────────────────────────────

def 初始化数据库(db路径: str):
    conn = sqlite3.connect(db路径)
    # 评论表：层级区分主评论/楼中楼，所属主评论记录楼中楼挂在哪条主评论下
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 评论 (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            uid       TEXT NOT NULL,
            昵称      TEXT,
            内容      TEXT,
            层级      TEXT,
            所属主评论 TEXT,
            抓取时间  INTEGER
        )
    """)
    # 待抓楼中楼表：主评论抓取时登记，rpid + 回复数
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 待抓楼中楼 (
            rpid    TEXT PRIMARY KEY,
            回复数  INTEGER,
            已完成  INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 进度 (
            键   TEXT PRIMARY KEY,
            值   TEXT
        )
    """)
    conn.commit()
    return conn


def 读取键值(conn, 键, 默认=""):
    cur = conn.execute("SELECT 值 FROM 进度 WHERE 键=?", (键,))
    row = cur.fetchone()
    return row[0] if row else 默认


def 写键值(conn, 键, 值):
    conn.execute("INSERT OR REPLACE INTO 进度 VALUES (?, ?)", (键, str(值)))
    conn.commit()


# 主评论进度
def 读取主进度(conn) -> tuple[int, int]:
    return int(读取键值(conn, "next_offset", 0)), int(读取键值(conn, "主评论总数", 0))

def 保存主进度(conn, next_offset: int, 总数: int):
    conn.execute("INSERT OR REPLACE INTO 进度 VALUES ('next_offset', ?)", (str(next_offset),))
    conn.execute("INSERT OR REPLACE INTO 进度 VALUES ('主评论总数', ?)", (str(总数),))
    conn.commit()


def 写入主评论(conn, 批次: list):
    """批次：[(uid, 昵称, 内容, rpid, rcount), ...]"""
    now = int(time.time())
    conn.executemany(
        "INSERT INTO 评论 (uid, 昵称, 内容, 层级, 所属主评论, 抓取时间) "
        "VALUES (?, ?, ?, '主评论', '', ?)",
        [(uid, 昵称, 内容, now) for uid, 昵称, 内容, _, _ in 批次]
    )
    # 登记待抓楼中楼（仅有回复的）
    conn.executemany(
        "INSERT OR IGNORE INTO 待抓楼中楼 (rpid, 回复数) VALUES (?, ?)",
        [(str(rpid), rcount) for _, _, _, rpid, rcount in 批次 if rcount > 0]
    )
    conn.commit()


def 写入楼中楼(conn, 所属rpid: str, 批次: list):
    """批次：[(uid, 昵称, 内容), ...]"""
    now = int(time.time())
    conn.executemany(
        "INSERT INTO 评论 (uid, 昵称, 内容, 层级, 所属主评论, 抓取时间) "
        "VALUES (?, ?, ?, '楼中楼', ?, ?)",
        [(uid, 昵称, 内容, 所属rpid, now) for uid, 昵称, 内容 in 批次]
    )
    conn.commit()


def 回复数分布(conn) -> dict:
    """统计待抓楼中楼表里各回复数门槛的评论条数"""
    分布 = {}
    for 门槛 in (1, 2, 5, 10, 20, 50, 100):
        cur = conn.execute("SELECT COUNT(*) FROM 待抓楼中楼 WHERE 回复数 >= ?", (门槛,))
        分布[门槛] = cur.fetchone()[0]
    return 分布


def 取待抓列表(conn, 阈值: int) -> list:
    """返回回复数 >= 阈值 且未完成的 (rpid, 回复数)"""
    cur = conn.execute(
        "SELECT rpid, 回复数 FROM 待抓楼中楼 WHERE 回复数 >= ? AND 已完成 = 0 ORDER BY 回复数 DESC",
        (阈值,)
    )
    return cur.fetchall()


def 标记完成(conn, rpid: str):
    conn.execute("UPDATE 待抓楼中楼 SET 已完成 = 1 WHERE rpid = ?", (rpid,))
    conn.commit()


def 从数据库读取结果(conn) -> dict:
    """聚合成 {uid: {"昵称": str, "主评论": [...], "楼中楼": [...]}}"""
    结果 = defaultdict(lambda: {"昵称": "", "主评论": [], "楼中楼": []})
    cur = conn.execute("SELECT uid, 昵称, 内容, 层级 FROM 评论 ORDER BY id")
    for uid, 昵称, 内容, 层级 in cur.fetchall():
        结果[uid]["昵称"] = 昵称 or ""
        if 层级 == "楼中楼":
            结果[uid]["楼中楼"].append(内容 or "")
        else:
            结果[uid]["主评论"].append(内容 or "")
    return 结果


# ── WBI 签名 ────────────────────────────────────────────────────────

def 获取混淆密钥(img_key: str, sub_key: str) -> str:
    raw = img_key + sub_key
    return "".join(raw[i] for i in MIXIN_KEY_ENC_TAB)[:32]


def wbi签名(params: dict, img_key: str, sub_key: str) -> dict:
    mixin_key = 获取混淆密钥(img_key, sub_key)
    params["wts"] = int(time.time())
    sorted_params = dict(sorted(params.items()))
    query = "&".join(
        f"{k}={re.sub(r'[!#$&+,/:;=?@\\[\\]]', '', str(v))}"
        for k, v in sorted_params.items()
    )
    params["w_rid"] = hashlib.md5((query + mixin_key).encode()).hexdigest()
    return params


async def 获取wbi密钥(凭证) -> tuple[str, str]:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.bilibili.com",
    }
    cookies = {
        "SESSDATA": 凭证.sessdata,
        "bili_jct": 凭证.bili_jct,
        "buvid3":   凭证.buvid3,
    }
    async with httpx.AsyncClient(headers=headers, cookies=cookies) as client:
        resp = await client.get("https://api.bilibili.com/x/web-interface/nav")
        data = resp.json()
    wbi_img = data["data"]["wbi_img"]
    def _key(url):
        return url.rsplit("/", 1)[-1].split(".")[0]
    return _key(wbi_img["img_url"]), _key(wbi_img["sub_url"])


# ── 抓取：主评论 ─────────────────────────────────────────────────────

def 构造请求头(凭证, referer):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Referer": referer,
    }
    cookies = {
        "SESSDATA": 凭证.sessdata,
        "bili_jct": 凭证.bili_jct,
        "buvid3":   凭证.buvid3,
    }
    return headers, cookies


async def 抓取主评论(cfg, conn, oid, referer, img_key, sub_key):
    凭证    = cfg["凭证"]
    评论类型 = cfg["评论类型"]
    最大条数 = cfg["最大条数"]
    间隔    = cfg["请求间隔秒"]

    next_offset, 总数 = 读取主进度(conn)
    if 总数 > 0:
        print(f"[主评论] 检测到上次进度：已抓 {总数} 条，从游标 {next_offset} 处续抓……")
    else:
        print(f"[主评论] 开始抓取……")

    限制说明 = f"最多 {最大条数} 条" if 最大条数 > 0 else "不限制，抓到底"
    print(f"[主评论] 抓取上限：{限制说明}")
    print("提示：随时按 Ctrl+C 可主动停止，进度自动保存。\n")

    headers, cookies = 构造请求头(凭证, referer)
    轮次 = 1

    async with httpx.AsyncClient(headers=headers, cookies=cookies, timeout=15) as client:
        while True:
            if 最大条数 > 0 and 总数 >= 最大条数:
                print(f"[主评论] 已达条数上限（{最大条数}），停止。")
                break

            params = wbi签名(
                {"oid": oid, "type": 评论类型, "mode": 3, "next": next_offset},
                img_key, sub_key,
            )
            try:
                resp = await client.get(
                    "https://api.bilibili.com/x/v2/reply/wbi/main", params=params)
                data = resp.json()
            except asyncio.CancelledError:
                raise
            except Exception as e:
                print(f"[主评论] 第 {轮次} 次请求失败：{e}，进度已保存。")
                break

            code = data.get("code", -1)
            if code != 0:
                print(f"[主评论] 接口错误码 {code}：{data.get('message','')}，进度已保存。")
                break

            cursor_info = data["data"].get("cursor", {})
            replies     = data["data"].get("replies") or []
            if not replies:
                print("[主评论] 已抓完全部主评论。")
                break

            本批 = []
            for r in replies:
                if 最大条数 > 0 and 总数 + len(本批) >= 最大条数:
                    break
                本批.append((
                    r["member"]["mid"],
                    清理非法字符(r["member"]["uname"]),
                    清理非法字符(r["content"]["message"]),
                    r["rpid"],
                    r.get("rcount", 0),
                ))
            写入主评论(conn, 本批)
            总数 += len(本批)

            next_offset = cursor_info.get("next", 0)
            is_end      = cursor_info.get("is_end", True)
            保存主进度(conn, next_offset, 总数)

            print(f"[主评论] 第 {轮次:>4} 次完成，本批 {len(本批)} 条，累计 {总数} 条（已落盘）……")
            轮次 += 1

            if is_end or (最大条数 > 0 and 总数 >= 最大条数):
                print("[主评论] 已到达末尾。")
                break
            await asyncio.sleep(间隔)

    return 总数


# ── 抓取：楼中楼 ─────────────────────────────────────────────────────

async def 抓取单条楼中楼(client, oid, 评论类型, root_rpid, img_key, sub_key, 间隔):
    """抓取某条主评论下的全部回复，返回 [(uid,昵称,内容),...]"""
    收集 = []
    pn = 1
    while True:
        params = wbi签名(
            {"oid": oid, "type": 评论类型, "root": root_rpid, "ps": 20, "pn": pn},
            img_key, sub_key,
        )
        resp = await client.get(
            "https://api.bilibili.com/x/v2/reply/reply", params=params)
        data = resp.json()
        if data.get("code", -1) != 0:
            break
        replies = data["data"].get("replies") or []
        if not replies:
            break
        for r in replies:
            收集.append((
                r["member"]["mid"],
                清理非法字符(r["member"]["uname"]),
                清理非法字符(r["content"]["message"]),
            ))
        # 楼中楼按 pn 翻页，replies 少于 20 说明到末页
        if len(replies) < 20:
            break
        pn += 1
        await asyncio.sleep(间隔)
    return 收集


async def 抓取所有楼中楼(cfg, conn, oid, referer, img_key, sub_key, 阈值):
    凭证    = cfg["凭证"]
    评论类型 = cfg["评论类型"]
    间隔    = cfg["请求间隔秒"]

    待抓 = 取待抓列表(conn, 阈值)
    if not 待抓:
        print(f"[楼中楼] 没有回复数 ≥ {阈值} 的评论需要抓取。")
        return 0

    print(f"[楼中楼] 共 {len(待抓)} 条主评论（回复数 ≥ {阈值}）需要抓取楼中楼……\n")

    headers, cookies = 构造请求头(凭证, referer)
    已处理 = 0
    新增回复 = 0

    async with httpx.AsyncClient(headers=headers, cookies=cookies, timeout=15) as client:
        for rpid, rcount in 待抓:
            try:
                回复列表 = await 抓取单条楼中楼(
                    client, oid, 评论类型, rpid, img_key, sub_key, 间隔)
            except asyncio.CancelledError:
                raise
            except Exception as e:
                print(f"[楼中楼] 评论 {rpid} 抓取失败：{e}，已保存进度，可续抓。")
                break

            写入楼中楼(conn, str(rpid), 回复列表)
            标记完成(conn, str(rpid))
            已处理 += 1
            新增回复 += len(回复列表)
            print(f"[楼中楼] ({已处理}/{len(待抓)}) 评论 {rpid}（回复数{rcount}）"
                  f"抓到 {len(回复列表)} 条，累计新增 {新增回复} 条……")
            await asyncio.sleep(间隔)

    print(f"\n[楼中楼] 本轮完成，新增 {新增回复} 条楼中楼评论。")
    return 新增回复


# ── 抓取总调度 ───────────────────────────────────────────────────────

async def 执行抓取(cfg, conn):
    凭证    = cfg["凭证"]
    评论类型 = cfg["评论类型"]

    if 评论类型 == 1:
        v = video.Video(bvid=cfg["bvid"], credential=凭证)
        oid = await v.get_aid() if asyncio.iscoroutinefunction(v.get_aid) else v.get_aid()
        referer = f"https://www.bilibili.com/video/{cfg['bvid']}"
        print(f"目标：视频 {cfg['bvid']}（aid={oid}）")
    else:
        oid = cfg["oid_直填"]
        referer = f"https://t.bilibili.com/{oid}"
        类型名 = "转发/纯文字动态" if 评论类型 == 17 else "图文相册动态"
        print(f"目标：{类型名}（oid={oid}，type={评论类型}）")

    print("正在获取 WBI 签名密钥……")
    img_key, sub_key = await 获取wbi密钥(凭证)

    阶段 = 读取键值(conn, "阶段", "main")

    try:
        # ── 阶段一：主评论 ──
        if 阶段 == "main":
            await 抓取主评论(cfg, conn, oid, referer, img_key, sub_key)
            写键值(conn, "阶段", "main_done")
            阶段 = "main_done"

        if not cfg["抓楼中楼"]:
            return

        # ── 阶段二：确定阈值（展示分布） ──
        阈值 = int(读取键值(conn, "回复阈值", -1))
        if 阈值 < 0:
            分布 = 回复数分布(conn)
            print("\n" + "=" * 50)
            print("主评论抓取完成，回复数分布如下（带回复的主评论）：")
            for 门槛, 数量 in 分布.items():
                print(f"  回复数 ≥ {门槛:>3} ：{数量} 条主评论")
            print("=" * 50)
            print("只有回复数 ≥ 阈值 的主评论才会抓取楼中楼。")
            print("填 1 = 抓所有带回复的（最完整但最慢）；建议默认 5。")
            while True:
                原始 = 询问("请输入楼中楼抓取阈值", 默认回复阈值)
                try:
                    阈值 = int(原始)
                    if 阈值 < 1:
                        print("  阈值至少为 1。")
                        continue
                    break
                except ValueError:
                    print("  请输入一个整数。")
            写键值(conn, "回复阈值", 阈值)
        else:
            print(f"\n[楼中楼] 沿用上次设定的阈值：{阈值}")

        # ── 阶段三：楼中楼 ──
        await 抓取所有楼中楼(cfg, conn, oid, referer, img_key, sub_key, 阈值)
        写键值(conn, "阶段", "all_done")

    except (asyncio.CancelledError, KeyboardInterrupt):
        print("\n⚑ 已手动停止，进度已保存。下次运行将从断点继续。\n")
        raise


# ── 报表 & 导出 ──────────────────────────────────────────────────────

def 生成报表(结果, 保留内容=True):
    """
    每个 UID 一行，区分主评论数 / 楼中楼数 / 总数。
    列：UID, 昵称, 主评论数, 楼中楼数, 总数, [评论内容]
    """
    行列表 = []
    for uid, 信息 in 结果.items():
        主 = 信息["主评论"]
        楼 = 信息["楼中楼"]
        总 = len(主) + len(楼)
        if 保留内容:
            块 = []
            if 主:
                块.append("【主评论】\n" + "\n".join(f"{i+1}. {c}" for i, c in enumerate(主)))
            if 楼:
                块.append("【楼中楼】\n" + "\n".join(f"{i+1}. {c}" for i, c in enumerate(楼)))
            内容 = 清理非法字符("\n\n".join(块))
            行列表.append([uid, 信息["昵称"], len(主), len(楼), 总, 内容])
        else:
            行列表.append([uid, 信息["昵称"], len(主), len(楼), 总])
    行列表.sort(key=lambda x: x[4], reverse=True)  # 按总数降序
    return 行列表


def 导出xlsx(行列表, 文件名, 保留内容=True):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    print("正在导出 Excel，请稍候……")
    wb = Workbook()
    ws = wb.active
    ws.title = "评论统计"

    if 保留内容:
        表头 = ["用户UID", "用户昵称", "主评论数", "楼中楼数", "总数", "评论内容"]
        宽度 = [16, 20, 10, 10, 8, 70]
    else:
        表头 = ["用户UID", "用户昵称", "主评论数", "楼中楼数", "总数"]
        宽度 = [16, 20, 10, 10, 8]

    ws.append(表头)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for 行 in 行列表:
        净行 = [清理非法字符(str(v)) if isinstance(v, str) else v for v in 行]
        ws.append(净行)

    for i, w in enumerate(宽度, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for 行 in ws.iter_rows(min_row=2):
        for cell in 行:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    路径 = f"{文件名}.xlsx"
    wb.save(路径)
    print(f"已导出 Excel：{路径}")


# ── 入口 ─────────────────────────────────────────────────────────────

async def 异步主流程():
    cfg = 收集配置()
    目标标识 = cfg["目标标识"]
    db路径 = f"评论缓存_{目标标识}.db"
    db是新的 = not os.path.exists(db路径)
    conn = 初始化数据库(db路径)

    if not db是新的:
        _, 已有 = 读取主进度(conn)
        if 已有 > 0:
            print(f"\n发现缓存文件 {db路径}，已存主评论 {已有} 条。")
            选择 = 询问("是否从上次中断处续抓？（y=续抓 / n=清空重新抓）", "y").lower()
            if 选择 == "n":
                conn.close()
                os.remove(db路径)
                conn = 初始化数据库(db路径)
                print("已清空缓存，重新开始。\n")

    被中断 = False
    try:
        await 执行抓取(cfg, conn)
    except (asyncio.CancelledError, KeyboardInterrupt):
        被中断 = True
    except Exception as e:
        print(f"\n抓取过程出错：{e}\n进度已保存，重新运行可续抓。")

    结果 = 从数据库读取结果(conn)
    conn.close()

    if not 结果:
        print("未抓取到任何评论，请检查 BV 号 / 动态ID / cookie / 网络。")
        return

    主评论数 = sum(len(v["主评论"]) for v in 结果.values())
    楼中楼数 = sum(len(v["楼中楼"]) for v in 结果.values())
    print("\n" + "=" * 50)
    print(f"当前数据库：主评论 {主评论数} 条，楼中楼 {楼中楼数} 条，"
          f"共 {主评论数+楼中楼数} 条，涉及 {len(结果)} 个用户。")
    print("=" * 50)

    if 被中断:
        导出选择 = 询问("已中断。是否用当前已抓数据导出 Excel？（y / n）", "y").lower()
        if 导出选择 != "y":
            print(f"已跳过导出，缓存保留在：{db路径}，下次运行可续抓。")
            return

    行列表 = 生成报表(结果, cfg["保留内容"])
    模式说明 = "完整保留评论内容" if cfg["保留内容"] else "仅统计评论数量"
    print(f"输出模式：{模式说明}")
    导出xlsx(行列表, cfg["输出文件名"], cfg["保留内容"])

    if not 被中断:
        删除 = 询问(f"\n是否删除缓存文件 {db路径}？（y=删除 / n=保留）", "y").lower()
        if 删除 == "y":
            os.remove(db路径)
            print("缓存已删除。")
        else:
            print(f"缓存已保留在：{db路径}")
    else:
        print(f"缓存保留在：{db路径}，下次运行可继续抓取剩余楼中楼。")


def main():
    try:
        asyncio.run(异步主流程())
    except Exception as e:
        print(f"\n程序出错：{e}")
    finally:
        input("\n按回车键退出……")


if __name__ == "__main__":
    main()
